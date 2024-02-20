from flask import Flask, render_template, request, send_file
from docxtpl import DocxTemplate
import os
import openpyxl
from docx2pdf import convert
import io
import subprocess
import sys
import docx2pdf
import convertapi


app = Flask(__name__)

#lấy đường dẫn của file đang thực thi
script_directory = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))

# Lấy đường dẫn đến tệp thực thi
exe_path = os.path.abspath(sys.argv[0])

# Lấy thư mục chứa tệp thực thi
exe_directory = os.path.dirname(exe_path)

# Xây dựng đường dẫn tới tập lệnh doc2pdf
doc2pdf_path = os.path.join(exe_directory, "doc2pdf")

with open('key.txt', 'r') as file:
    content = file.read()

convertapi.api_secret = f'{content}'


@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        name1 = request.form["name1"]
        name2 = request.form["name2"]
        phone = request.form["phone"]
        cvtv = request.form["CVTV"]
        strain = request.form["strain"]
        tfrc = request.form["tfrc"]
        tfrcLH = request.form["tfrcLH"]

        thisistfrc = tfrcFunction(tfrcLH)
        tinhcach, tuongtac, pphoc = general(strain)

        ageOfCus = request.form["age"]
        tiemnang1 = request.form["tiemnang1"]
        tiemnang2 = request.form["tiemnang2"]
        tiemnang3 = request.form["tiemnang3"]
        tiemnang4 = request.form["tiemnang4"]
        tiemnang9 = request.form["tiemnang9"]
        tiemnang10 = request.form["tiemnang10"]
        vakChoice = request.form["veaka"]

        ttmvtroi1 = request.form["ttmvt1"]
        ttmvtroi2 = request.form["ttmvt2"]
        ttmvtroi3 = request.form["ttmvt3"]
        ttmvtroi4 = request.form["ttmvt4"]

        tn1, kn1 = tiemnangVuotTroi(tiemnang1, ageOfCus)
        tn2, kn2 = tiemnangVuotTroi(tiemnang2, ageOfCus)
        tn3, kn3 = tiemnangVuotTroi(tiemnang3, ageOfCus)
        tn4, kn4 = tiemnangVuotTroi(tiemnang4, ageOfCus)
        tn9, kn9 = tiemnangKemVuotTroi(tiemnang9, ageOfCus)
        tn10, kn10 = tiemnangKemVuotTroi(tiemnang10, ageOfCus)
        vakDetail = chosenVAK(vakChoice)

        chisoEQ, chisoAQ = chisoEQAQ(ageOfCus)

        ttmdetail1, ttmgen1 = trithongMinh(ttmvtroi1)
        ttmdetail2, ttmgen2 = trithongMinh(ttmvtroi2)
        ttmdetail3, ttmgen3 = trithongMinh(ttmvtroi3)
        ttmdetail4, ttmgen4 = trithongMinh(ttmvtroi4)

        prioritized = priority()
        priorField = getPriorityField(prioritized)

        # Load the template Word document using docxtpl
        template_path = "template.docx"
        doc = DocxTemplate(template_path)

        # Define context data for rendering
        context = {
            "name1": name1.upper(),
            "name2": name2.upper(),
            "phone": phone,
            "CVTV": cvtv.upper(),
            "CVTVfoot": capitalize_name(cvtv),
            "type": strain.upper(),
            "tfrc": tfrc,
            "tinhcach": tinhcach,
            "tuongtac": tuongtac,
            "pphoc": pphoc,
            "thisistfrc": thisistfrc,
            "tiemnang1": tn1,
            "khuyennghi1": kn1,
            "tiemnang2": tn2,
            "khuyennghi2": kn2,
            "tiemnang3": tn3,
            "khuyennghi3": kn3,
            "tiemnang4": tn4,
            "khuyennghi4": kn4,
            "tiemnang9": tn9,
            "khuyennghi9": kn9,
            "tiemnang10": tn10,
            "khuyennghi10": kn10,
            "veaka": vakDetail,
            "chisoeq": chisoEQ,
            "chisoaq": chisoAQ,
            "ttmgen1": ttmgen1,
            "ttmdetail1": ttmdetail1,
            "ttmgen2": ttmgen2,
            "ttmdetail2": ttmdetail2,
            "ttmgen3": ttmgen3,
            "ttmdetail3": ttmdetail3,
            "ttmgen4": ttmgen4,
            "ttmdetail4": ttmdetail4,
            "recommend": priorField,
        }

        # Render the template with context data
        doc.render(context)

        # Save the rendered document
        output_path = f"{script_directory}/BÁO CÁO TÓM TẮT {name2.upper()}.docx"
        doc.save(output_path)

        conversion_type = request.form["conversion_type"]
        if conversion_type == "pdf":
            try:
                pdf_output_path = (
                    f"{script_directory}/BÁO CÁO TÓM TẮT {name2.upper()}.pdf"
                )
                # Convert DOCX to PDF using subprocess
                # subprocess.run(["docx2pdf", output_path, pdf_output_path], shell=True)
                try:
                    result = convertapi.convert('pdf', { 'File': output_path })

                    # save to file
                    result.file.save(pdf_output_path)
                    
                    #subprocess.run(["docx2pdf", output_path, pdf_output_path], shell=False)
                except Exception as e:
                    return str(e)
                # Prepare PDF file for download
                #pdf_output = f"{script_directory}/BÁO CÁO TÓM TẮT {name2.upper()}.pdf"
                #pdf_output = pdf_output_path
                
                pdf_data = open(pdf_output_path, "rb").read()

                # Provide the PDF file for download
                return send_file(
                    io.BytesIO(pdf_data),
                    mimetype="application/pdf",
                    download_name=f"BÁO CÁO TÓM TẮT {name2.upper()}.pdf",
                    as_attachment=True,
                )
            except Exception as e:
                return str(e)
        else:
            return send_file(output_path, as_attachment=True)
    else:
        return render_template("index.html")


def capitalize_name(name):
    words = name.split()
    capitalized_words = [word.capitalize() for word in words]
    capitalized_name = " ".join(capitalized_words)
    return capitalized_name


def get_excel_data(num_cell):
    excel_path = "metadata.xlsx"  # Đường dẫn tới tệp Excel của bạn
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook["TK Trẻ em"]

    # Lấy dữ liệu từ ô A1
    cell_value = sheet[num_cell].value

    # Lấy ghi chú (note) của ô A1
    cell_note = sheet[num_cell].comment
    cell_note_text = cell_note.text if cell_note else ""

    workbook.close()
    return cell_note_text


def get_data_type(num_cell):
    excel_path = "metadata.xlsx"  # Đường dẫn tới tệp Excel của bạn
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook["TK Trẻ em"]

    # Lấy dữ liệu từ ô A1
    cell_value = sheet[num_cell].value

    workbook.close()
    return str(cell_value)


def general(strain):
    for i in range(5, 22):
        if strain == get_data_type(f"A{i}").lower().replace(" ", ""):
            tinhcach = get_excel_data("B" + str(i))
            tuongtac = get_excel_data("C" + str(i))
            pphoc = get_excel_data("D" + str(i))
            return tinhcach, tuongtac, pphoc


def tfrcFunction(tfrcVal):
    if tfrcVal == "veryhigh":
        tfrcFuc = get_excel_data("E5")
        return tfrcFuc
    elif tfrcVal == "high":
        tfrcFuc = get_excel_data("E6")
        return tfrcFuc
    elif tfrcVal == "good":
        tfrcFuc = get_excel_data("E7")
        return tfrcFuc
    elif tfrcVal == "medArch":
        tfrcFuc = get_excel_data("E8")
        return tfrcFuc
    elif tfrcVal == "med":
        tfrcFuc = get_excel_data("E9")
        return tfrcFuc


valuesTN = [
    "quanly",
    "lanhdao",
    "logic",
    "tuongtuong",
    "vandongtinh",
    "vandongtho",
    "ngonngu",
    "amthanh",
    "quansat",
    "thammy",
]


def tiemnangVuotTroi(tiemnangVal, tuoi):
    for i in range(0, len(valuesTN)):
        x = i + 5
        if tuoi == "dotuoi1":
            if tiemnangVal == valuesTN[i]:
                tiemnang = get_excel_data(f"F{x}")
                tiemnangDetail = get_excel_data(f"G{x}")
                return tiemnang, tiemnangDetail
        elif tuoi == "dotuoi2":
            if tiemnangVal == valuesTN[i]:
                tiemnang = get_excel_data(f"F{x}")
                tiemnangDetail = get_excel_data(f"I{x}")
                return tiemnang, tiemnangDetail


def tiemnangKemVuotTroi(tiemnangVal, tuoi):
    for i in range(0, len(valuesTN)):
        x = i + 5
        if tuoi == "dotuoi1":
            if tiemnangVal == valuesTN[i]:
                tiemnang = get_excel_data(f"F{x}")
                tiemnangDetail = get_excel_data(f"H{x}")
                return tiemnang, tiemnangDetail
        elif tuoi == "dotuoi2":
            if tiemnangVal == valuesTN[i]:
                tiemnang = get_excel_data(f"F{x}")
                tiemnangDetail = get_excel_data(f"J{x}")
                return tiemnang, tiemnangDetail


def chosenVAK(vakChoice):
    vakArr = ["avk", "vak", "kva", "akv", "kav", "vka"]
    for i in range(0, len(vakArr)):
        x = i + 5
        if vakChoice == vakArr[i]:
            vakDetail = get_excel_data(f"K{x}")
            return vakDetail


def chisoEQAQ(age):
    if age == "dotuoi1":
        chisoEQ = get_excel_data("L5")
        chisoAQ = get_excel_data("L6")
        return chisoEQ, chisoAQ
    elif age == "dotuoi2":
        chisoEQ = get_excel_data("M5")
        chisoAQ = get_excel_data("M6")
        return chisoEQ, chisoAQ


ttmvtVal = [
    "languagex",
    "logicx",
    "khonggianx",
    "vandongx",
    "amnhacx",
    "noitamx",
    "tuongtacx",
    "thiennhienx",
    "hiensinhx",
]


def trithongMinh(ttmvt):
    for i in range(0, len(ttmvtVal)):
        x = i + 5
        if ttmvt == ttmvtVal[i]:
            thisTrithongminh = get_excel_data(f"N{x}")
            thisGeneralTTM = get_data_type(f"N{x}")
            return thisTrithongminh, thisGeneralTTM


def priority():
    priority_fields = {
        "Chính trị-Luật": int(request.form.get("field1")),
        "Truyền thông": int(request.form.get("field2")),
        "Giáo dục": int(request.form.get("field3")),
        "Quản trị-Điều hành": int(request.form.get("field4")),
        "Tâm lý học": int(request.form.get("field5")),
        "Ngôn ngữ-Đối ngoại": int(request.form.get("field6")),
        "Xã hội học": int(request.form.get("field7")),
        "Văn học-Sử học-Triết học": int(request.form.get("field8")),
        "Nghệ thuật-Giải trí": int(request.form.get("field9")),
        "Y khoa": int(request.form.get("field10")),
        "Thiết kế": int(request.form.get("field11")),
        "Công nghệ thông tin": int(request.form.get("field12")),
        "Kỹ sư": int(request.form.get("field13")),
        "Tài chính-Kế toán": int(request.form.get("field14")),
        "Khách sạn-Du lịch": int(request.form.get("field15")),
        "Kiến trúc": int(request.form.get("field16")),
        "Thể thao": int(request.form.get("field17")),
        "Quân đội": int(request.form.get("field18")),
        "Kinh doanh bán hàng-Tiếp thị": int(request.form.get("field19")),
    }
    sorted_fields = sorted(priority_fields, key=priority_fields.get)
    prioritized_fields = [
        field for field in sorted_fields if priority_fields[field] > 0
    ]
    return prioritized_fields


def getPriorityField(prioritized_fields):
    priorFields = [
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
    ]
    result = ""
    for i in range(0, len(prioritized_fields)):
        for j in range(1, 20):
            if str(prioritized_fields[i]) == get_data_type(f"Q{j}"):
                priorFields[i] = get_excel_data(f"Q{j}")
    for x in range(0, len(prioritized_fields)):
        result += priorFields[x] + "\n"
    return result


##############################################################################################################################
@app.route("/adult", methods=["GET", "POST"])
def adult():
    if request.method == "POST":
        name1 = request.form["name1"]
        name2 = request.form["name2"]
        phone = request.form["phone"]
        cvtv = request.form["CVTV"]
        strain = request.form["strain"]
        tfrc = request.form["tfrc"]
        tfrcLH = request.form["tfrcLH"]

        tiemnang1 = request.form["tiemnang1"]
        tiemnang2 = request.form["tiemnang2"]
        tiemnang3 = request.form["tiemnang3"]
        tiemnang4 = request.form["tiemnang4"]
        tiemnang9 = request.form["tiemnang9"]
        tiemnang10 = request.form["tiemnang10"]
        vakChoice = request.form["veaka"]

        ttmvtroi1 = request.form["ttmvt1"]
        ttmvtroi2 = request.form["ttmvt2"]
        ttmvtroi3 = request.form["ttmvt3"]
        ttmvtroi4 = request.form["ttmvt4"]

        thisistfrc = tfrcFunction_Adult(tfrcLH)
        tinhcach, khuyennghi = general_Adult(strain)

        tn1, kn1 = tiemnangVuotTroi_Adult(tiemnang1)
        tn2, kn2 = tiemnangVuotTroi_Adult(tiemnang2)
        tn3, kn3 = tiemnangVuotTroi_Adult(tiemnang3)
        tn4, kn4 = tiemnangVuotTroi_Adult(tiemnang4)
        tn9, kn9 = tiemnangKemVuotTroi_Adult(tiemnang9)
        tn10, kn10 = tiemnangKemVuotTroi_Adult(tiemnang10)

        vakDetail = chosenVAK_Adult(vakChoice)

        ttmdetail1, ttmgen1 = trithongMinh_Adult(ttmvtroi1)
        ttmdetail2, ttmgen2 = trithongMinh_Adult(ttmvtroi2)
        ttmdetail3, ttmgen3 = trithongMinh_Adult(ttmvtroi3)
        ttmdetail4, ttmgen4 = trithongMinh_Adult(ttmvtroi4)

        prioritized_Adult = priority_Adult()
        priorField_Adult = getPriorityField_Adult(prioritized_Adult)

        # Load the template Word document using docxtpl
        template_path = "templateAdult.docx"
        doc = DocxTemplate(template_path)

        # Define context data for rendering
        context = {
            "name1": name1.upper(),
            "name2": name2.upper(),
            "phone": phone,
            "CVTV": cvtv.upper(),
            "CVTVfoot": capitalize_name(cvtv),
            "type": strain.upper(),
            "tfrc": tfrc,
            "tinhcach": tinhcach,
            "khuyennghi": khuyennghi,
            "thisistfrc": thisistfrc,
            "tiemnang1": tn1,
            "khuyennghi1": kn1,
            "tiemnang2": tn2,
            "khuyennghi2": kn2,
            "tiemnang3": tn3,
            "khuyennghi3": kn3,
            "tiemnang4": tn4,
            "khuyennghi4": kn4,
            "tiemnang9": tn9,
            "khuyennghi9": kn9,
            "tiemnang10": tn10,
            "khuyennghi10": kn10,
            "veaka": vakDetail,
            "ttmgen1": ttmgen1,
            "ttmdetail1": ttmdetail1,
            "ttmgen2": ttmgen2,
            "ttmdetail2": ttmdetail2,
            "ttmgen3": ttmgen3,
            "ttmdetail3": ttmdetail3,
            "ttmgen4": ttmgen4,
            "ttmdetail4": ttmdetail4,
            "recommend": priorField_Adult,
        }

        # Render the template with context data
        doc.render(context)
        # Save the rendered document
        output_path = f"{script_directory}/BÁO CÁO TÓM TẮT {name2.upper()}.docx"
        doc.save(output_path)

        conversion_type = request.form["conversion_type"]

        if conversion_type == "pdf":
            try:
                pdf_output_path = (
                    f"{script_directory}/BÁO CÁO TÓM TẮT {name2.upper()}.pdf"
                )
                
                result = convertapi.convert('pdf', { 'File': output_path })

                # save to file
                result.file.save(pdf_output_path)
                
                # Convert DOCX to PDF using subprocess
                # subprocess.run(["docx2pdf", output_path, pdf_output_path], shell=True)
                #subprocess.run([doc2pdf_path, output_path], shell=False)
                # Prepare PDF file for download
                
                pdf_data = open(pdf_output_path, "rb").read()

                # Provide the PDF file for download
                return send_file(
                    io.BytesIO(pdf_data),
                    mimetype="application/pdf",
                    download_name=f"BÁO CÁO TÓM TẮT {name2.upper()}.pdf",
                    as_attachment=True,
                )
            except Exception as e:
                return str(e)
        else:
            return send_file(output_path, as_attachment=True)
    return render_template("adult.html")


def capitalize_name(name):
    words = name.split()
    capitalized_words = [word.capitalize() for word in words]
    capitalized_name = " ".join(capitalized_words)
    return capitalized_name


def get_excel_data_Adult(num_cell):
    excel_path = "metadata.xlsx"  # Đường dẫn tới tệp Excel của bạn
    workbook = openpyxl.load_workbook(excel_path)
    # Chọn sheet theo tên
    sheet = workbook["TK Người lớn"]
    # Lấy dữ liệu từ ô A1
    cell_value = sheet[num_cell].value
    # Lấy ghi chú (note) của ô A1
    cell_note = sheet[num_cell].comment
    cell_note_text = cell_note.text if cell_note else ""
    workbook.close()
    return cell_note_text


# pyinstaller --onefile --add-data "templates;templates" --add-data "template.docx;template.docx" --add-data "templateAdult.docx;templateAdult.docx" --add-data "metadata.xlsx;metadata.xlsx" app.py


def get_data_type_Adult(num_cell):
    excel_path = "metadata.xlsx"  # Đường dẫn tới tệp Excel của bạn
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook["TK Người lớn"]
    # Lấy dữ liệu từ ô A1
    cell_value = sheet[num_cell].value
    workbook.close()
    return str(cell_value)


def general_Adult(strain):
    for i in range(4, 19):
        if strain == get_data_type_Adult(f"A{i}").lower().replace(" ", ""):
            tinhcach = get_excel_data_Adult("B" + str(i))
            khuyennghi = get_excel_data_Adult("C" + str(i))
            return tinhcach, khuyennghi


def tfrcFunction_Adult(tfrcVal):
    if tfrcVal == "veryhigh":
        tfrcFuc = get_excel_data_Adult("D4")
        return tfrcFuc
    elif tfrcVal == "high":
        tfrcFuc = get_excel_data_Adult("D5")
        return tfrcFuc
    elif tfrcVal == "good":
        tfrcFuc = get_excel_data_Adult("D6")
        return tfrcFuc
    elif tfrcVal == "medArch":
        tfrcFuc = get_excel_data_Adult("D7")
        return tfrcFuc
    elif tfrcVal == "med":
        tfrcFuc = get_excel_data_Adult("D8")
        return tfrcFuc


valuesTN_Adult = [
    "quanly",
    "lanhdao",
    "logic",
    "tuongtuong",
    "vandongtinh",
    "vandongtho",
    "ngonngu",
    "amthanh",
    "quansat",
    "thammy",
]


def tiemnangVuotTroi_Adult(tiemnangVal):
    for i in range(0, len(valuesTN_Adult)):
        x = i + 4
        if tiemnangVal == valuesTN_Adult[i]:
            tiemnang = get_excel_data_Adult(f"E{x}")
            tiemnangDetail = get_excel_data_Adult(f"F{x}")
            return tiemnang, tiemnangDetail


def tiemnangKemVuotTroi_Adult(tiemnangVal):
    for i in range(0, len(valuesTN_Adult)):
        x = i + 4
        if tiemnangVal == valuesTN_Adult[i]:
            tiemnang = get_excel_data_Adult(f"E{x}")
            tiemnangDetail = get_excel_data_Adult(f"G{x}")
            return tiemnang, tiemnangDetail


def chosenVAK_Adult(vakChoice):
    vakArr = ["avk", "vak", "kva", "akv", "kav", "vka"]
    for i in range(0, len(vakArr)):
        x = i + 4
        if vakChoice == vakArr[i]:
            vakDetail = get_excel_data_Adult(f"H{x}")
            return vakDetail


ttmvtVal_Adult = [
    "languagex",
    "logicx",
    "khonggianx",
    "vandongx",
    "amnhacx",
    "noitamx",
    "tuongtacx",
    "thiennhienx",
    "hiensinhx",
]


def trithongMinh_Adult(ttmvt):
    for i in range(0, len(ttmvtVal_Adult)):
        x = i + 4
        if ttmvt == ttmvtVal_Adult[i]:
            thisTrithongminh = get_excel_data_Adult(f"I{x}")
            thisGeneralTTM = get_data_type_Adult(f"I{x}")
            return thisTrithongminh, thisGeneralTTM


def priority_Adult():
    priority_fields = {
        "Chính trị-Luật": int(request.form.get("field1")),
        "Truyền thông": int(request.form.get("field2")),
        "Giáo dục": int(request.form.get("field3")),
        "Quản trị-Điều hành": int(request.form.get("field4")),
        "Tâm lý học": int(request.form.get("field5")),
        "Ngôn ngữ-Đối ngoại": int(request.form.get("field6")),
        "Xã hội học": int(request.form.get("field7")),
        "Văn học-Sử học-Triết học": int(request.form.get("field8")),
        "Nghệ thuật-Giải trí": int(request.form.get("field9")),
        "Y khoa": int(request.form.get("field10")),
        "Thiết kế": int(request.form.get("field11")),
        "Công nghệ thông tin": int(request.form.get("field12")),
        "Kỹ sư": int(request.form.get("field13")),
        "Tài chính-Kế toán": int(request.form.get("field14")),
        "Khách sạn-Du lịch": int(request.form.get("field15")),
        "Kiến trúc": int(request.form.get("field16")),
        "Thể thao": int(request.form.get("field17")),
        "Quân đội": int(request.form.get("field18")),
        "Kinh doanh bán hàng-Tiếp thị": int(request.form.get("field19")),
    }
    sorted_fields = sorted(priority_fields, key=priority_fields.get)
    prioritized_fields = [
        field for field in sorted_fields if priority_fields[field] > 0
    ]
    return prioritized_fields


def getPriorityField_Adult(prioritized_fields):
    priorFields = [
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
    ]
    result = ""
    for i in range(0, len(prioritized_fields)):
        for j in range(1, 20):
            if str(prioritized_fields[i]) == get_data_type_Adult(f"Q{j}"):
                priorFields[i] = get_excel_data_Adult(f"Q{j}")
    for x in range(0, len(prioritized_fields)):
        result += priorFields[x] + "\n"
    return result


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=6001, debug=True)
